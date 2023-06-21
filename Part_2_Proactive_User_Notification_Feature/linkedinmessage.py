import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

linkedin_username = "anonym*******@gmail.com"
linkedin_password = "Tes*****"

excel_file = "linkedin_data.xlsx"

driver = webdriver.Edge()

def login_to_linkedin(username, password):
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)

    username_field = driver.find_element(By.ID, "username")
    username_field.send_keys(username)
    password_field = driver.find_element(By.ID, "password")
    password_field.send_keys(password)
    password_field.send_keys(Keys.RETURN)

    try:
        WebDriverWait(driver, 10).until(EC.url_contains("https://www.linkedin.com/feed/"))
    except TimeoutException:
        print("Login failed. Please check your credentials.")
        driver.quit()
        exit()

def get_unread_data():
    driver.get("https://www.linkedin.com/feed/")
    driver.implicitly_wait(10)  

    navbar = driver.find_element(By.CSS_SELECTOR, ".global-nav__content")
    unread_counts = navbar.find_elements(By.CSS_SELECTOR, ".notification-badge__count")

    message_count = int(unread_counts[0].text)
    notification_count = int(unread_counts[1].text)

    return message_count, notification_count


def save_data_to_excel(messages, notifications):
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="Date")
        sheet.cell(row=1, column=2, value="Time")
        sheet.cell(row=1, column=3, value="Unread Messages")
        sheet.cell(row=1, column=4, value="Unread Notifications")
    else:
        workbook = load_workbook(excel_file)
        sheet = workbook.active

    last_row = sheet.max_row
    current_datetime = datetime.now()
    sheet.cell(row=last_row + 1, column=1, value=current_datetime.date())
    sheet.cell(row=last_row + 1, column=2, value=current_datetime.time())
    sheet.cell(row=last_row + 1, column=3, value=messages)
    sheet.cell(row=last_row + 1, column=4, value=notifications)
    workbook.save(excel_file)

def send_email_notification(messages, notifications, prev_messages=None, prev_notifications=None):
    # Email configuration
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "anonym*********@gmail.com"
    sender_password = "nsdgd*********"
    recipient_email = "gowthams*********@gmail.com"

    email_subject = "LinkedIn Unread Notifications"

    email_body = f"""
        <html>
        <head>
            <style>
            body {{
                font-family: Arial, sans-serif;
                background-color: #f5f5f5;
            }}
            .card {{
                padding: 20px;
                border-radius: 4px;
                box-shadow: 0px 0px 5px 2px rgba(0, 0, 0, 0.1);
                background-color: white;
            }}
            h1, h2 {{
                color: #00b300;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                padding: 8px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }}
            th {{
                font-weight: bold;
                font-size: 16px;
            }}
            td {{
                font-size: 14px;
            }}
            .highlight {{
                font-weight: bold;
                color: #00b300;
            }}
            </style>
        </head>
        <body>
            <div class="card">
            <h1>LinkedIn Unread Notifications</h1>
            <table>
                <tr>
                    <th>Date</th>
                    <th>Time</th>
                    <th>Unread Messages</th>
                    <th>Unread Notifications</th>
                </tr>
                <tr>
                    <td>{datetime.now().date()}</td>
                    <td>{datetime.now().time()}</td>
                    <td>{messages}</td>
                    <td>{notifications}</td>
                </tr>
            </table>
            """

    if prev_messages is not None and prev_notifications is not None:
        email_body += f"""
            <h2>Comparison with past data:</h2>
            <p><span class="highlight">Previous Unread Messages :</span> {messages - prev_messages}</p>
            <p><span class="highlight">Previous Unread Notifications :</span> {notifications - prev_notifications}</p>
        """

    email_body += """
            </div>
        </body>
        </html>
    """

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = email_subject
    message.attach(MIMEText(email_body, "html"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)
            print("Email notification sent successfully.")
    except Exception as e:
        print(f"Failed to send email notification. Error: {str(e)}")


def main():
    login_to_linkedin(linkedin_username, linkedin_password)

    while True:
        prev_messages, prev_notifications = get_unread_data()
        time.sleep(3 * 60 * 60)  
        current_messages, current_notifications = get_unread_data()
        save_data_to_excel(current_messages, current_notifications)
        send_email_notification(current_messages, current_notifications, prev_messages, prev_notifications)

if __name__ == "__main__":
    main()
